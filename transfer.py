import openpyxl, csv, pprint

#* This program transfers student data from my main gradebook file to the school's template in the proper format. It also writes all the threes
# TODO: Make the program transfer attendance, too

#* This loads all student data from the gradebook template provided by the school
rostersheet = openpyxl.load_workbook('roster.xlsx')
class_section = ""

#* This creates a list of all student ID numbers used in the grade book, in order to avoid adding grades for students who have dropped out
id_list = []
for row in rostersheet.active:
    id_list.append(int(row[0].value))


#? This didn't need to be a function, and also can't really be made smaller without restructuring it entirely. This file is just meant as a rough tool anyway
def write_threes():

    #* This loads the grade data from my personal gradebook and initializes a dictionary where the useful data will be stored later
    gradebook = openpyxl.load_workbook('2022S2FinalGradesMain.xlsx')
    classdata_dict = {}

    #* This iterates through the class sections I have grades for in my gradebook
    for section in gradebook.sheetnames: 
        section_num = [section[0], section[1:]]

        #* This transforms the class section name into a more appropriate format for .xlsx files
        section_key = "-".join(section_num)

        #* This initializes a list in which dictionaries of individual student data will be stored, accessible by the class section key
        classdata_dict.setdefault(section_key, [])

        #* This establishes which row the following process will be looking at
        cursor = 2

        #* Initialize id_num as 0 to begin the while loop
        id_num = 0
        while id_num != None:

            #* Initialize a dictionary of individual student data which will be added to the list associated with the current class section
            student_dict = {}

            #* Get the ID number in the current row, and break the loop if it's out of the range of values
            id_num = gradebook[section][f"B{cursor}"].value
            if id_num == None:
                maxes[section_key] = cursor - 1
                break
            else:

                #* Check if the ID number from my gradebook is still in the school-provided list, and if so, collect all relevant information and put it into the dictionary
                if int(id_num) in id_list:
                    id_num = int(id_num)
                    student_dict["id"] = id_num
                    student_dict["premid_1"] = gradebook[section][f"D{cursor}"].value
                    student_dict["premid_2"] = gradebook[section][f"E{cursor}"].value
                    student_dict["premid_3"] = gradebook[section][f"F{cursor}"].value
                    student_dict["premid_total"] = gradebook[section][f"G{cursor}"].value
                    student_dict["midterm"] = gradebook[section][f"I{cursor}"].value
                    student_dict["postmid_1"] = gradebook[section][f"M{cursor}"].value
                    student_dict["postmid_2"] = gradebook[section][f"N{cursor}"].value
                    student_dict["postmid_3"] = gradebook[section][f"O{cursor}"].value
                    student_dict["postmid_total"] = gradebook[section][f"P{cursor}"].value
                    student_dict["final"] = gradebook[section][f"T{cursor}"].value
                    classdata_dict[section_key].append(student_dict)
            cursor += 1
        

    #* Iterate through each student in the roster sheet
    for student in rostersheet.active:

        #* Then iterate through all students from the gradebook data just extracted
        for section in classdata_dict.keys():
            for student_import in classdata_dict[section]:

                #* For each student, get their index number, gender, name, and surname from the school-provided list
                if student_import["id"] == student[0].value:
                    student_import["index"] = student[5].value
                    student_import["gender"] = student[1].value
                    student_import["name"] = student[2].value
                    student_import["surname"] = student[3].value
    
    #* Iterate through each class section I have grades for again
    for section in gradebook.sheetnames:
        section_key = "-".join([section[0], section[1:]])

        #* Open the template provided by the school
        target = openpyxl.load_workbook("test.xlsx")

        #* Open the sheet where the threes will be written
        sheet = target["threes"]

        #* Open the sheet where grades will be written
        grade_sheet = target["grades"]

        #* Establish the page heading in the correct format, then add it to both pages
        page_heading =  f"รหัสวิชา..................................วิชา.................................................................                         ห้อง  {' / '.join([section[0], section[1:]])}     ภาคเรียนที่  2 /  2565"
        sheet["A2"] = page_heading
        sheet["A28"] = page_heading
        sheet["A54"] = page_heading
        grade_sheet["A2"] = page_heading
        grade_sheet["A28"] = page_heading
        grade_sheet["A54"] = page_heading

        #* Iterate through each student in the class data for this class section
        for student in classdata_dict[section_key]:

            #* Make sure the cursor is pointed at the correct row for the format of each page
            if student["index"] <= 20:
                row_num = student["index"] + 6
            elif 20 < student["index"] <= 40:
                row_num = student["index"] + 12
            else:
                row_num = student["index"] + 18

            #* Start adding values to the fields
            sheet[f"A{row_num}"] = student["index"]
            sheet[f"B{row_num}"] = student["id"]
            sheet[f"C{row_num}"] = student["gender"]
            sheet[f"D{row_num}"] = student["name"]

            #* Make a list of each column the threes will go into, then iterate through each one, adding the threes
            threes  = ["I", "L", "N", "Q", "T", "W", "AA", "AD", "AE", "AK", "AQ", "AW", "BC", "BI", "BJ"]
            for col in threes:
                sheet[f"{col}{row_num}"] = 3      

            #* Adjust the cursor for larger classes on the gradebook page
            if student["index"] > 40:
                row_num = student["index"] + 21

            #* Add all the values to the proper cells
            grade_sheet[f"A{row_num}"] = student["index"]
            grade_sheet[f"B{row_num}"] = student["id"]
            grade_sheet[f"C{row_num}"] = student["gender"]
            grade_sheet[f"D{row_num}"] = student["name"]
            grade_sheet[f"E{row_num}"] = student["surname"]
            grade_sheet[f"F{row_num}"] = student["premid_1"]
            grade_sheet[f"G{row_num}"] = student["premid_2"]
            grade_sheet[f"H{row_num}"] = student["premid_3"]
            grade_sheet[f"J{row_num}"] = student["premid_total"]
            grade_sheet[f"X{row_num}"] = student["midterm"]
            grade_sheet[f"AA{row_num}"] = student["postmid_1"]
            grade_sheet[f"AB{row_num}"] = student["postmid_2"]
            grade_sheet[f"AC{row_num}"] = student["postmid_3"]
            grade_sheet[f"AE{row_num}"] = student["postmid_total"]
            grade_sheet[f"AP{row_num}"] = int(student["premid_total"]) + int(student["midterm"]) + int(student["postmid_total"])
            grade_sheet[f"AT{row_num}"] = student["final"]
            grade_sheet[f"AU{row_num}"] = int(grade_sheet[f"AP{row_num}"].value) + int(student["final"])

        #* Save the file
        target.save(f"{section_key}SGS.xlsx")

        #* Notify the user that the current section has been completed
        print(f"finished {section_key}")


write_threes()